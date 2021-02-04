# -*- coding: utf-8 -*-
"""
Created on Mon Nov  9 16:50:27 2020

@author: Xu Mingjun
"""

import torch
import numpy as np
from openpyxl import load_workbook
from pendigits_train_save import *
    
#python 函数入口
if __name__ == "__main__":
    # 设置随机化种子（和训练保持一致）
    np.random.seed(2)
    
    #读取表格中的数据
    workbook = load_workbook('dataset\pendigits.xlsx')
    booksheet = workbook.get_sheet_by_name('Sheet1')
    
    #初始化Ndarray用于存放表中变量
    data = np.zeros((10992,17))
    
    #读取sheet1中的数据到numpy的矩阵data中
    for i in range(10992):    
        for j in range(17): 
            data[i][j] = booksheet.cell(row=i+2, column=j+1).value
    
    #默认对Ndarray的第0维打乱（对样本的顺序进行打乱）
    np.random.shuffle(data)  
          
    #分出992个测试数据
    test_label = data[10000:10992,-1]
    test_feature = data[10000:10992,0:-1]
    
    #对label进行onehot编码（多分类）
    test_label_oh = np.eye(10)[test_label.astype(int)]
    
    #对feature归一化处理
    test_feature = test_feature / test_feature.max(axis=0)
    
    #训练和测试进行类型转换  
    test_feature_t = torch.from_numpy(test_feature)
    test_label_t = torch.from_numpy(test_label_oh)
    
    test_feature_t_f = torch.tensor(test_feature_t,dtype=torch.float32)
    test_label_t_f = torch.tensor(test_label_t,dtype=torch.float32)
    
    #######################测试过程##############################
    
    model = torch.load('./SaveData/model.pt', map_location='cpu')
    
    model.eval()    #保证BN和dropout不发生变化
    
    cnt = 0 #初始化正确的计数值
    
    #输入训练集得到测试结果
    test_out = model(test_feature_t_f)
    _, test_out_np= torch.max(test_out,1)   #onehot解码，返回值第一个是最大值（不需要），第二个是最大值的序号

    #迭代922个测试样本输出和统计    
    for test_i in range(992):
    
        print("No.",test_i,"\npre:",test_out_np.numpy()[test_i],"\nGT:",test_label[test_i])
        print("****************")
        if test_out_np.numpy()[test_i] == test_label[test_i]:
            #print("correct")
            cnt += 1
    
    #正确率计算
    correct_rate = cnt/992.0
    print("correct_rate:",correct_rate)
